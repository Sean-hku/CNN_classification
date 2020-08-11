from train_result.config import task_folder, batch_folder, folder_path
import openpyxl
import os

model_txt = os.path.join(folder_path, batch_folder + ".csv")

models = []
with open(model_txt, "r") as f:
    for line in f.readlines():
        try:
            models.append(int(line.split(" ")[-1][:-2]))
        except:
            continue

# print(models)

trained_folder = "../result/{}-{}".format(task_folder, batch_folder)

import os
file_trained = [int(file) for file in os.listdir(trained_folder) if len(file)<=4]

# print(sorted(trained))
rest_file = [item for item in models if item not in file_trained]
print("Not in files:")
print(rest_file)

train_log_name = "alphapose_aic/result/aic_origin_result.xlsx"
wb = openpyxl.load_workbook(train_log_name)
sheet_names = wb.get_sheet_names()
ws = wb.get_sheet_by_name(sheet_names[0])
logs = []
for row in range(ws.max_row-1):
    if ws.cell(row+2,1).value is not None:
        logs.append(ws.cell(row+2,1).value)
wb.close()
print("Not in logs:")
rest_log = [item for item in models if item not in logs]
print(rest_log)

rest_all = [item for item in rest_file if item in rest_log]
print("Not in both:")
print(rest_all)
